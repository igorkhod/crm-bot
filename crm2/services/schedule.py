# === Автогенерированный заголовок: crm2/services/schedule.py
# Список верхнеуровневых объектов файла (классы и функции).
# Обновляется вручную при изменении состава функций/классов.
# Классы: Session
# Функции: _norm, _parse_date, _find_header_row, _cohort_id_from_filename, _load_one_file, load_all, get_user_cohort_id, upcoming, format_next, format_sessions_brief, next_training_text_for_user
# === Конец автозаголовка
#
# === Файл: crm2/services/schedule.py
# Аннотация: модуль CRM, расписание и события. Внутри классы: Session, функции: _norm, _parse_date, _find_header_row, _cohort_id_from_filename, _load_one_file....
# Добавлено автоматически 2025-08-21 05:43:17

# crm2/services/schedule.py
"""
Загрузка расписаний из .xlsx для потоков и форматирование выдачи.

Форматы колонок (регистр не важен):
- start_date | дата начала | start | startdate
- end_date   | дата окончания | end | enddate
- topic_code | индекс | code
- title      | тема | название
- annotation | аннотация | описание

Парсер сам ищет строку с заголовками среди первых 10 строк,
поэтому можно оставлять декоративную шапку (например, "расписание 2025 2 поток")
над таблицей — это не сломает загрузку.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import re
from pathlib import Path
from typing import Iterable, Optional, Dict, List, Tuple

from openpyxl import load_workbook

from crm2.services.users import get_user_by_telegram, get_user_cohort_id_by_tg
from crm2.db.core import get_db_connection
import sqlite3


# Где лежат файлы расписаний
DATA_DIR = Path(__file__).resolve().parents[1] / "data"

# Принимаемые названия колонок (в любом регистре)
START_HEADERS = {"start_date", "дата начала", "start", "startdate", "начало"}
END_HEADERS = {"end_date", "дата окончания", "end", "enddate", "окончание"}
CODE_HEADERS = {"topic_code", "code", "индекс"}
TITLE_HEADERS = {"title", "тема", "название"}
ANN_HEADERS = {"annotation", "аннотация", "описание"}

HEADER_SCAN_ROWS = 10  # сколько верхних строк просканировать в поисках хедера


@dataclass
class Session:
    start: date
    end: date
    code: str = ""
    title: str = ""
    annotation: str = ""


# ---------- утилиты ----------

def _norm(s) -> str:
    return str(s).strip().lower() if s is not None else ""

def _parse_date(cell_val) -> date:
    """Поддержка datetime/date и строковых форматов 'YYYY-MM-DD' или 'DD.MM.YYYY'."""
    if isinstance(cell_val, datetime):
        return cell_val.date()
    if isinstance(cell_val, date):
        return cell_val
    s = _norm(cell_val)
    if not s:
        raise ValueError("Пустая дата")
    # ISO 'YYYY-MM-DD'
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        pass
    # RU 'DD.MM.YYYY'
    try:
        return datetime.strptime(s, "%d.%m.%Y").date()
    except ValueError:
        pass
    # Последняя попытка — fromisoformat (бывает '2025-09-13 00:00:00')
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        raise ValueError(f"Не смог распарсить дату: {cell_val!r}")


def _find_header_row(ws) -> Tuple[int, Dict[str, int]]:
    """
    Возвращает (row_index, mapping), где mapping — словарь с индексами нужных колонок.
    Ищем среди первых HEADER_SCAN_ROWS строк.
    """
    # основной проход — ищем строку с заголовками
    for r in range(1, min(ws.max_row, HEADER_SCAN_ROWS) + 1):
        # values_only=True -> получаем кортеж значений, НЕ ячейки
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
        headers = [_norm(x) for x in row]

        def _find_idx(candidates: Iterable[str]) -> Optional[int]:
            for i, h in enumerate(headers):
                if h in candidates:
                    return i
            return None

        i_start = _find_idx(START_HEADERS)
        i_end = _find_idx(END_HEADERS)
        i_code = _find_idx(CODE_HEADERS)
        i_title = _find_idx(TITLE_HEADERS)
        i_ann = _find_idx(ANN_HEADERS)

        if i_start is not None and i_end is not None:
            return r, {
                "start": i_start,
                "end": i_end,
                "code": i_code,
                "title": i_title,
                "annotation": i_ann,
            }

    # для понятного сообщения об ошибке сформируем превью первых строк
    preview = []
    for r in range(1, min(ws.max_row, HEADER_SCAN_ROWS) + 1):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
        preview.append([_norm(x) for x in row])

    raise ValueError(
        "В xlsx не найдены нужные столбцы среди первых строк.\n"
        f"Ожидались: один из {START_HEADERS} и один из {END_HEADERS}\n"
        f"Первые строки: {preview}"
    )



def _cohort_id_from_filename(path: Path) -> Optional[int]:
    """
    Вытаскиваем id потока из имени файла вида '... 1 поток.xlsx' или '... 2 поток.xlsx'.
    Если числа нет — вернём None.
    """
    m = re.search(r"(\d+)\s*поток", path.stem.lower())
    return int(m.group(1)) if m else None


def _load_one_file(path: Path) -> List[Session]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header_row_idx, cols = _find_header_row(ws)
    sessions: List[Session] = []

    # Идём по строкам после хедера
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        values = list(row)
        if all(v is None for v in values):
            continue  # пустая строка

        try:
            start = _parse_date(values[cols["start"]])
            end = _parse_date(values[cols["end"]])

            code = values[cols["code"]] if cols["code"] is not None else ""
            title = values[cols["title"]] if cols["title"] is not None else ""
            ann = values[cols["annotation"]] if cols["annotation"] is not None else ""

            sessions.append(
                Session(
                    start=start,
                    end=end,
                    code=str(code or "").strip(),
                    title=str(title or "").strip(),
                    annotation=str(ann or "").strip(),
                )
            )
        except Exception:
            # пропускаем строку, если даты кривые
            continue

    sessions.sort(key=lambda s: (s.start, s.end))
    return sessions


# ---------- публичный API, который зовут хендлеры ----------

def load_all() -> Dict[int, List[Session]]:
    """
    Читает все файлы вида 'расписание *.xlsx' из crm2/data и
    возвращает {cohort_id: [Session, ...]}.
    """
    result: Dict[int, List[Session]] = {}
    has_files = False
    for path in DATA_DIR.glob("расписание *.xlsx"):
        has_files = True
        cohort_id = _cohort_id_from_filename(path)
        if not cohort_id:
            continue
        result[cohort_id] = _load_one_file(path)
    if has_files:
        return result
    # --- Fallback: берём из БД таблицу sessions (если xlsx отсутствуют) ---
    with get_db_connection() as con:
        con.row_factory = sqlite3.Row
        cur = con.execute("""
            SELECT id, start_date, COALESCE(end_date, start_date) AS end_date,
                   COALESCE(topic_code,'') AS topic_code,
                   COALESCE(title,'')      AS title,
                   COALESCE(annotation,'') AS annotation,
                   cohort_id
            FROM sessions
            ORDER BY start_date ASC, id ASC
        """)
        rows = cur.fetchall()
    for r in rows:
        try:
            s = Session(
                start = datetime.fromisoformat(r["start_date"]).date(),
                end = datetime.fromisoformat(r["end_date"]).date(),
                code = str(r["topic_code"] or ""),
                title = str(r["title"] or ""),
                annotation = str(r["annotation"] or ""),
            )
        except Exception:
            continue
        cid = int(r["cohort_id"]) if r["cohort_id"] is not None else -1
        result.setdefault(cid, []).append(s)
    for v in result.values():
        v.sort(key=lambda s: (s.start, s.end))
    return result

def get_user_cohort_id(telegram_id: int) -> Optional[int]:
    # Сначала пробуем явный cohort_id в users
    u = get_user_by_telegram(telegram_id)
    if u and u.get("cohort_id"):
        return u["cohort_id"]
    # Совместимость: если поток хранится в participants — берём оттуда
    return get_user_cohort_id_by_tg(telegram_id)


def upcoming(telegram_id: int, *, limit: int = 1) -> List[Session]:
    """Ближайшие (не прошедшие) занятия.
    Если у пользователя не задан cohort_id — показываем ближайшие по всем потокам."""
    cohort_id = get_user_cohort_id(telegram_id)
    today = date.today()
    all_by_cohort = load_all()

    if cohort_id is not None and cohort_id in all_by_cohort:
        pool = all_by_cohort[cohort_id]
    else:
        # без потока — берём всё
        pool = [s for items in all_by_cohort.values() for s in items]

    return [s for s in pool if s.end >= today][:limit]

def format_next(s: Session) -> str:
    return f"Ближайшее занятие: {s.start.strftime('%d.%m.%Y')} — {s.end.strftime('%d.%m.%Y')}"


def format_sessions_brief(sessions: List[Session]) -> str:
    if not sessions:
        return "Расписание занятий:\n• пока пусто"
    lines = [f"• {s.start.strftime('%d.%m.%Y')} — {s.end.strftime('%d.%m.%Y')}" for s in sessions]
    return "Расписание занятий:\n" + "\n".join(lines)


def next_training_text_for_user(telegram_id: int) -> str:
    """Текст «ближайшего занятия» (или пустая строка, если нет данных/потока)."""
    items = upcoming(telegram_id, limit=1)
    return format_next(items[0]) if items else ""