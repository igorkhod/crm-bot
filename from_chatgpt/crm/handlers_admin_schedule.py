from aiogram import Router, types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
import aiosqlite
from datetime import datetime, timedelta

from from_chatgpt.crm.db import DB_PATH


# Роутер админ-раздела «Расписание»
admin_schedule_router = Router()


# --------------------------
# Машина состояний (FSM)
# --------------------------
class AddSessionFSM(StatesGroup):
    choose_cohort = State()    # выбор существующего потока
    create_cohort = State()    # ожидание ввода нового cohort_id
    enter_seq = State()        # номер сессии
    choose_kind = State()      # ПН/ПТГ
    enter_date = State()       # дата первого дня dd.mm.yyyy
    enter_title = State()      # название сессии


# --------------------------
# Стартовое мини-меню админа
# --------------------------
@admin_schedule_router.message(Command("админ_расписание"))
async def admin_menu(message: types.Message):
    """
    Показываем только два действия по требованию:
    – «Выбрать поток»
    – «Создать поток»
    """
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, keyboard=[
        [types.KeyboardButton(text="📚 Выбрать поток")],
        [types.KeyboardButton(text="➕ Создать поток")],
    ])
    await message.answer("Админ-панель: выберите или создайте поток.", reply_markup=kb)


# ---------------------------------------
# 1) «Выбрать поток»: таблица + кнопки
# ---------------------------------------
@admin_schedule_router.message(F.text == "📚 Выбрать поток")
async def choose_cohort(message: types.Message, state: FSMContext):
    """
    Выводим текстовую таблицу «№ | ближайшая дата» по всем потокам,
    у которых есть будущие занятия, и строим inline-кнопки 1..N.
    """
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        rows = await (await db.execute(
            """
            SELECT cohort_id, MIN(date) AS next_date
            FROM session_days
            WHERE date >= DATE('now')
            GROUP BY cohort_id
            ORDER BY next_date
            """
        )).fetchall()

    if not rows:
        await message.answer("Пока нет потоков с будущими занятиями.")
        return

    # «Таблица» в тексте
    lines = ["№ | ближайшая дата"]
    buttons = []
    for i, r in enumerate(rows, start=1):
        d = datetime.fromisoformat(r["next_date"]).strftime("%d.%m.%Y")
        lines.append(f"{i} | {d}  [{r['cohort_id']}]")
        buttons.append([InlineKeyboardButton(text=str(i),
                                             callback_data=f"pick_cohort:{r['cohort_id']}")])

    await message.answer(
        "\n".join(lines),
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )
    # ждём «логический» выбор потока (необязательное состояние, но пусть будет)
    await state.set_state(AddSessionFSM.choose_cohort)


# ---------------------------------------
# Callback: показать полное расписание
# ---------------------------------------
@admin_schedule_router.callback_query(F.data.startswith("pick_cohort:"))
async def cb_pick_cohort(cb: types.CallbackQuery, state: FSMContext):
    """
    По нажатию на номер из таблицы запоминаем поток и
    показываем расписание этого потока полностью.
    """
    cohort_id = cb.data.split(":", 1)[1]
    await state.update_data(cohort_id=cohort_id)

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        rows = await (await db.execute(
            """
            SELECT seq, course_kind, title, date_start, date_end
            FROM course_sessions
            WHERE cohort_id = ?
            ORDER BY seq, course_kind
            """, (cohort_id,)
        )).fetchall()

    if not rows:
        await cb.message.edit_text(f"Поток ‘{cohort_id}’ выбран. Расписание пока пусто.")
        await cb.answer()
        return

    out = [f"🗂 Расписание потока {cohort_id}:"]
    for r in rows:
        d1 = datetime.fromisoformat(r["date_start"]).strftime("%d.%m.%Y")
        d2 = datetime.fromisoformat(r["date_end"]).strftime("%d.%m.%Y")
        out.append(f"{r['seq']} {r['course_kind']} — {r['title']} ({d1}–{d2})")

    await cb.message.edit_text("\n".join(out))
    await cb.answer()


# ---------------------------------------
# 2) «Создать поток»: ввод cohort_id
# ---------------------------------------
@admin_schedule_router.message(F.text == "➕ Создать поток")
async def create_cohort(message: types.Message, state: FSMContext):
    """Просим ввести идентификатор нового потока и переходим в состояние create_cohort."""
    await message.answer("Введите идентификатор потока (например, recruitmen_2025_03):")
    await state.set_state(AddSessionFSM.create_cohort)


@admin_schedule_router.message(AddSessionFSM.create_cohort, F.text.regexp(r"^[A-Za-z0-9_]{4,}$"))
async def save_new_cohort(message: types.Message, state: FSMContext):
    """
    Принимаем cohort_id ТОЛЬКО в состоянии create_cohort, чтобы не перехватывать
    никнеймы при логине и прочие сообщения.
    """
    cohort_id = message.text.strip()
    await state.update_data(cohort_id=cohort_id)
    await message.answer(
        f"Поток ‘{cohort_id}’ готов к заполнению. "
        f"Можете добавлять сессии («➕ Добавить сессию») или импортировать из Excel."
    )


# ---------------------------------------
# 3) Добавление сессии (пошагово)
# ---------------------------------------
@admin_schedule_router.message(F.text == "➕ Добавить сессию")
async def add_session_start(message: types.Message, state: FSMContext):
    data = await state.get_data()
    cohort_id = data.get("cohort_id")
    if not cohort_id:
        await message.answer("Сначала выберите поток или создайте новый.")
        return
    await state.set_state(AddSessionFSM.enter_seq)
    await message.answer("Введите номер сессии (целое число, например: 10):")


@admin_schedule_router.message(AddSessionFSM.enter_seq, F.text.regexp(r"^\d{1,2}$"))
async def add_session_seq(message: types.Message, state: FSMContext):
    await state.update_data(seq=int(message.text.strip()))
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, keyboard=[
        [types.KeyboardButton(text="ПН"), types.KeyboardButton(text="ПТГ")]
    ])
    await state.set_state(AddSessionFSM.choose_kind)
    await message.answer("Выберите тип сессии: ПН или ПТГ", reply_markup=kb)


@admin_schedule_router.message(AddSessionFSM.choose_kind, F.text.in_({"ПН", "ПТГ"}))
async def add_session_kind(message: types.Message, state: FSMContext):
    await state.update_data(kind=message.text.strip())
    await state.set_state(AddSessionFSM.enter_date)
    await message.answer("Введите дату первого дня в формате ДД.ММ.ГГГГ (например, 20.09.2025):")


@admin_schedule_router.message(AddSessionFSM.enter_date, F.text.regexp(r"^\d{2}\.\d{2}\.\d{4}$"))
async def add_session_date(message: types.Message, state: FSMContext):
    d1 = datetime.strptime(message.text.strip(), "%d.%m.%Y")
    d2 = d1 + timedelta(days=1)  # второй день автоматически
    await state.update_data(date_start=d1.date().isoformat(), date_end=d2.date().isoformat())
    await state.set_state(AddSessionFSM.enter_title)
    await message.answer("Введите название сессии (или «-», чтобы поставить по умолчанию):")


@admin_schedule_router.message(AddSessionFSM.enter_title)
async def add_session_finish(message: types.Message, state: FSMContext):
    data = await state.get_data()
    title = message.text.strip()
    if title == "-":
        title = ("Психонетика " if data["kind"] == "ПН" else "Психотехнологии гармонии ") + str(data["seq"])

    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            """
            INSERT OR IGNORE INTO course_sessions
                (cohort_id, seq, course_kind, title, date_start, date_end)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (data["cohort_id"], data["seq"], data["kind"], title, data["date_start"], data["date_end"]),
        )
        await db.commit()

    await state.clear()
    await message.answer("✅ Сессия добавлена. Ещё? Нажмите «➕ Добавить сессию» или используйте импорт из Excel.")


# ---------------------------------------
# 4) Импорт расписания из Excel (как на скрине: дата1 | дата2 | "N ПН/ПТГ")
# ---------------------------------------
@admin_schedule_router.message(F.text == "📥 Импорт из Excel")
async def import_excel_prompt(message: types.Message):
    await message.answer(
        "Пришлите .xlsx. Формат строк: ДД.ММ.ГГГГ | ДД.ММ.ГГГГ | 'N ПН/ПТГ'. "
        "cohort_id берём из выбранного потока."
    )


@admin_schedule_router.message(F.document)
async def import_excel_file(message: types.Message, state: FSMContext):
    doc = message.document
    if not doc.file_name.lower().endswith(".xlsx"):
        return  # молча игнорируем другие файлы

    data = await state.get_data()
    cohort_id = data.get("cohort_id")
    if not cohort_id:
        await message.answer("Сначала выберите поток («📚 Выбрать поток») или создайте новый.")
        return

    file = await message.bot.get_file(doc.file_id)
    tmp_path = f"/tmp/{doc.file_name}"
    await message.bot.download_file(file.file_path, destination=tmp_path)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(tmp_path)
        ws = wb.active
    except Exception as e:
        await message.answer(f"Не удалось прочитать файл: {e}")
        return

    inserted = 0
    async with aiosqlite.connect(DB_PATH) as db:
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # пропускаем строки, где первая ячейка не дата (например «набор март 2025»)
            c1 = row[0]
            if isinstance(c1, str):
                try:
                    d1 = datetime.strptime(c1.strip(), "%d.%m.%Y")
                except Exception:
                    continue
            else:
                d1 = c1
            c2 = row[1]
            if isinstance(c2, str):
                try:
                    d2 = datetime.strptime(c2.strip(), "%d.%m.%Y")
                except Exception:
                    continue
            else:
                d2 = c2

            label = str(row[2]).strip() if row[2] is not None else ""
            if " " not in label:
                continue
            num_str, kind = label.split(None, 1)
            try:
                seq = int(num_str)
            except Exception:
                continue
            kind = kind.strip().upper()
            if kind not in {"ПН", "ПТГ"}:
                continue

            title = ("Психонетика " if kind == "ПН" else "Психотехнологии гармонии ") + str(seq)

            await db.execute(
                """
                INSERT OR IGNORE INTO course_sessions
                    (cohort_id, seq, course_kind, title, date_start, date_end)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (cohort_id, seq, kind, title, d1.date().isoformat(), d2.date().isoformat()),
            )
            inserted += 1
        await db.commit()

    await message.answer(f"✅ Импорт завершён. Добавлено строк: {inserted}.")
